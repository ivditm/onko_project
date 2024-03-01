from commands import (table_06_ins,
                      table_pop_data_insert,
                      table_state_cancer_insert,
                      table_state_help_57_insert,
                      table_57_insert,
                      table_07_insert)


from os import listdir
from os.path import isfile, join


from openpyxl import load_workbook
import re
import sqlite3


data_path = 'C:/Dev/onko_project/data/'
pattern = r'(\b\w+\b) \((С[^\)]+)\)'
fragment = """
СВЕДЕНИЯ О ЛЕЧЕНИИ ЗЛОКАЧЕСТВЕННЫХ НОВООБРАЗОВАНИЙ (ЗНО), ВПЕРВЫЕ
ЗАРЕГИСТРИРОВАННЫХ В 2021 Г., ПОДЛЕЖАЩИХ РАДИКАЛЬНОМУ ЛЕЧЕНИЮ
"""
fragment_1 = """
СВЕДЕНИЯ О  КОНТИНГЕНТЕ БОЛЬНЫХ СО ЗЛОКАЧЕСТВЕННЫМИ НОВООБРАЗОВАНИЯМИ,
 СОСТОЯЩЕМ НА УЧЕТЕ В ОНКОЛОГИЧЕСКИХ УЧРЕЖДЕНИЯХ В 2021 Г.
"""
conn = sqlite3.connect('data.db')


onlyfiles = [f for f in listdir(data_path) if isfile(join(data_path, f))]
cr_12_59 = [
    join(data_path, f_name) for f_name in onlyfiles
    if (12 <= int(f_name.split('_')[2]) <= 59)
    and 'Злокачественные' in f_name.split('_')]
cr_69_103 = [
    join(data_path, f_name) for f_name in onlyfiles
    if 69 <= int(f_name.split('_')[2]) <= 103
    and 'Злокачественные' in f_name.split('_')]
help_57_85 = [
    join(data_path, f_name) for f_name in onlyfiles
    if 58 <= int(f_name.split('_')[2]) <= 85
    and 'Состояние' in f_name.split('_')]
help_24_53 = [
    join(data_path, f_name) for f_name in onlyfiles
    if 24 <= int(f_name.split('_')[2]) <= 53
    and 'Состояние' in f_name.split('_')]
others = [
    join(data_path, f_name) for f_name in onlyfiles
    if join(data_path, f_name) not in cr_12_59
    and join(data_path, f_name) not in cr_69_103
    and join(data_path, f_name) not in help_57_85
    and join(data_path, f_name) not in help_24_53
]


def fast_test(l1, l2, l3, l4):
    assert not set(l1) & set(l2) & set(l3) & set(l4), 'fast test failed'


def load_first_list(l1, conn, table):
    cursor = conn.cursor()
    try:
        cursor.execute('BEGIN TRANSACTION')
        for f_name in l1:
            active_sheet = load_workbook(f_name).active
            table_name = active_sheet['B4'].value
            i = 8
            while active_sheet[f"A{i}"].value is not None:
                cursor.execute(table_pop_data_insert.format(table=table), (
                        active_sheet[f'A{i}'].value,
                        table_name,
                        active_sheet[f'B{i}'].value,
                        active_sheet[f'C{i}'].value,
                        active_sheet[f'D{i}'].value,
                        active_sheet[f'E{i}'].value,
                        active_sheet[f'F{i}'].value,
                        active_sheet[f'G{i}'].value,
                        active_sheet[f'H{i}'].value,
                        active_sheet[f'I{i}'].value,
                        active_sheet[f'J{i}'].value,
                        active_sheet[f'K{i}'].value,
                        active_sheet[f'L{i}'].value,
                        active_sheet[f'M{i}'].value
                        ))
                i += 1
            print(active_sheet["A1"].value, table_name, 'done')
        conn.commit()
    except Exception as e:
        print(f"Error processing file: {e}")
        conn.rollback()


def load_second_fille(l2, conn):
    cursor = conn.cursor()
    try:
        cursor.execute('BEGIN TRANSACTION')
        for f_name in l2[1:]:
            active_sheet = load_workbook(f_name).active
            i = 4
            table_name = active_sheet["A1"].value
            table_name = table_name.replace(fragment, '')
            match = re.search(r'(.*?) (Таблица.*)', table_name)
            table_name = match.group(1).strip()
            while active_sheet[f"A{i}"].value is not None:
                cursor.execute(table_state_cancer_insert,
                               (table_name,
                                active_sheet[f"A{i}"].value,
                                active_sheet[f"B{i}"].value,
                                active_sheet[f"C{i}"].value,
                                active_sheet[f"D{i}"].value,
                                active_sheet[f"E{i}"].value,
                                active_sheet[f"F{i}"].value,
                                active_sheet[f"G{i}"].value,
                                active_sheet[f"H{i}"].value,
                                active_sheet[f"I{i}"].value))
                i += 1
            print(active_sheet["A1"].value, 'done')
        conn.commit()
    except sqlite3.Error as e:
        print(f"An error occurred: {e}")
        conn.rollback()


def load_th_fille(l3, conn):
    cursor = conn.cursor()
    try:
        cursor.execute('BEGIN TRANSACTION')
        for f_name in l3:
            wb = load_workbook(f_name, read_only=True)
            i = 4
            table_name = wb.active["A1"].value
            table_name = table_name.replace(fragment_1, '')
            match = re.search(r'(.*?) (Таблица.*)', table_name)
            table_name = match.group(1).strip()
            sheet_names = wb.sheetnames
            if len(sheet_names) > 1:
                first_sheet = wb[sheet_names[0]]
                second_sheet = wb[sheet_names[1]]
                while first_sheet[f"A{i}"].value is not None:
                    cursor.execute(table_state_help_57_insert,
                                   (table_name,
                                    first_sheet[f"A{i}"].value,
                                    first_sheet[f"B{i}"].value,
                                    first_sheet[f"C{i}"].value,
                                    first_sheet[f"D{i}"].value,
                                    first_sheet[f"E{i}"].value,
                                    first_sheet[f"F{i}"].value,
                                    first_sheet[f"G{i}"].value,
                                    first_sheet[f"H{i}"].value,
                                    first_sheet[f"I{i}"].value,
                                    second_sheet[f"B{i+1}"].value,
                                    second_sheet[f"C{i+1}"].value,
                                    second_sheet[f"D{i+1}"].value,
                                    second_sheet[f"E{i+1}"].value,
                                    second_sheet[f"F{i+1}"].value,
                                    second_sheet[f"G{i+1}"].value,
                                    second_sheet[f"H{i+1}"].value,
                                    second_sheet[f"I{i+1}"].value))
                    i += 1
                print(first_sheet["A1"].value, 'done')
            else:
                k = 4
                while first_sheet[f"A{k}"].value is not None:
                    cursor.execute(table_state_help_57_insert,
                                   (table_name,
                                    first_sheet[f"A{k}"].value,
                                    first_sheet[f"B{k}"].value,
                                    first_sheet[f"C{k}"].value,
                                    first_sheet[f"D{k}"].value,
                                    first_sheet[f"E{k}"].value,
                                    first_sheet[f"F{k}"].value,
                                    first_sheet[f"G{k}"].value,
                                    first_sheet[f"H{k}"].value,
                                    first_sheet[f"I{k}"].value,
                                    0, 0, 0, 0, 0, 0, 0, 0))
                    k += 1
        conn.commit()
    except sqlite3.Error as e:
        print(f"An error occurred: {e}")
        conn.rollback()


def insert_table_6(conn):
    cursor = conn.cursor()
    try:
        cursor.execute('BEGIN TRANSACTION')
        active_sheet = load_workbook(
            'C:/Dev/onko_project/'
            'data/2021_Таблица_006_'
            'Злокачественные_новообразования_в_РФ_'
            '(заболеваемость_и_смертность).xlsx').active
        i = 5
        while active_sheet[f"A{i}"].value is not None:
            cursor.execute(table_06_ins, (
                                active_sheet[f"A{i}"].value,
                                active_sheet[f"B{i}"].value,
                                active_sheet[f"C{i}"].value,
                                active_sheet[f"D{i}"].value,
                                active_sheet[f"E{i}"].value,
                                active_sheet[f"F{i}"].value,
                                active_sheet[f"G{i}"].value,
                                active_sheet[f"H{i}"].value,
                                active_sheet[f"I{i}"].value,
                                active_sheet[f'J{i}'].value,
                                active_sheet[f'K{i}'].value,
                                active_sheet[f'L{i}'].value,
                                active_sheet[f'M{i}'].value))
            i += 1
        conn.commit()
    except sqlite3.Error as e:
        print(f"An error occurred: {e}")
        conn.rollback()


def insert_table_57(conn):
    cursor = conn.cursor()
    try:
        cursor.execute('BEGIN TRANSACTION')
        active_sheet = load_workbook(
            'C:/Dev/onko_project/data'
            '/2021_Таблица_057_Состояние_онко_помощи_в_РФ.xlsx').active
        i = 6
        while active_sheet[f"A{i}"].value is not None:
            cursor.execute(table_57_insert, (
                                active_sheet[f"A{i}"].value,
                                active_sheet[f"B{i}"].value,
                                active_sheet[f"C{i}"].value,
                                active_sheet[f"D{i}"].value,
                                active_sheet[f"E{i}"].value,
                                active_sheet[f"F{i}"].value,
                                active_sheet[f"G{i}"].value,
                                active_sheet[f"H{i}"].value,
                                active_sheet[f"I{i}"].value,
                                active_sheet[f'J{i}'].value,
                                active_sheet[f'K{i}'].value,
                                active_sheet[f'L{i}'].value,
                                active_sheet[f'M{i}'].value,
                                active_sheet[f"N{i}"].value,
                                active_sheet[f"O{i}"].value,
                                active_sheet[f"P{i}"].value,
                                active_sheet[f"Q{i}"].value,
                                active_sheet[f"R{i}"].value,
                                active_sheet[f"S{i}"].value,
                                active_sheet[f"T{i}"].value,
                                active_sheet[f"U{i}"].value,
                                active_sheet[f"V{i}"].value,
                                active_sheet[f'W{i}'].value))
            i += 1
        conn.commit()
    except sqlite3.Error as e:
        print(f"An error occurred: {e}")
        conn.rollback()


def insert_07(conn):
    cursor = conn.cursor()
    try:
        cursor.execute('BEGIN TRANSACTION')
        active_sheet = load_workbook('C:/Dev/onko_project/data'
                                     '/2021_Таблица_007_'
                                     'Злокачественные_новообразования'
                                     '_в_РФ_(заболеваемость_'
                                     'и_смертность).xlsx').active
        i = 5
        while active_sheet[f"A{i}"].value is not None:
            cursor.execute(table_07_insert, (
                        active_sheet[f'A{i}'].value,
                        active_sheet[f'B{i}'].value,
                        active_sheet[f'C{i}'].value,
                        active_sheet[f'D{i}'].value,
                        active_sheet[f'E{i}'].value,
                        active_sheet[f'F{i}'].value,
                        active_sheet[f'G{i}'].value
                        ))
            i += 1
        conn.commit()
    except Exception as e:
        print(f"Error processing file: {e}")
        conn.rollback()


def main():
    try:
        load_first_list(cr_12_59, conn, 'population_data')
        print('first_table done')
        load_first_list(cr_69_103, conn, 'mort_cancer')
        print('second_table done')
        load_second_fille(help_57_85, conn)
        print('второй список done')
        load_th_fille(help_24_53, conn)
        print('третий список done')
        insert_table_6(conn)
        print('table 6 done')
        insert_07(conn)
        print('table 7 done')
        insert_table_57(conn)
        print('table 57 done')
    except sqlite3.Error as e:
        print(f"An error occurred: {e}")
        conn.rollback()
    finally:
        conn.close()


if __name__ == '__main__':
    fast_test(cr_12_59, help_57_85, help_24_53, others)
    main()
